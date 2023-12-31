from flask import jsonify, request, Blueprint, send_file, make_response
import requests
import json
from russiannames.parser import NamesParser
import sqlite3
from openpyxl import Workbook
import io

from config import bearer


fiosplitter = Blueprint('fiosplitter', __name__)


DB_NAME = 'tools4u.db'

headers = {
    'Content-Type': 'application/json',
    'Authorization' : f'{bearer}',
    'Accept': 'application/vnd.yclients.v2+json'
}



def create_table():
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    cursor.execute('CREATE TABLE IF NOT EXISTS fiosplitter (id INTEGER PRIMARY KEY AUTOINCREMENT, salon_id INTEGER , data TEXT)')
    conn.commit()
    conn.close()
create_table()

def safeClients(salon,data):
    try:
        conn = sqlite3.connect(DB_NAME)
        cursor = conn.cursor()
        clients = str(data)
        # clients = json.dumps(data)
        cursor.execute("""
                   INSERT INTO fiosplitter (salon_id, data)
                   VALUES (?, ?)
               """, (int(salon), clients))
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"Error {e}")

clientsdata = []
newdata = []
errors = []

@fiosplitter.route('/getClients', methods=['POST'])
def getcli():
    try:
        response = request.json
        salon = response["salon_id"]
        login = response["login"]
        password = response["password"]

        if password == "":
            token = login
            clientsdata = getClients(salon, headers, token)
        else:
            usertoken = getToken(login,password)
            clientsdata = getClients(salon, headers, usertoken)
        return jsonify({"status": "success",  "text": clientsdata})
    except Exception as e:
        return jsonify({'status': 'error', 'text': f'{e}'})


def getToken(login, password):
    url = f"https://api.yclients.com/api/v1/auth"
    payload = json.dumps({
        "login": f"{login}",
        "password": f"{password}"
    })
    response = requests.request("POST", url, headers=headers, data=payload).json()
    if response["success"] == True:
        return response["data"]["user_token"]
    else:
        return f"Ошибка. {response['meta']}"
def getClients(salon, headers, token):
    page = 1
    global clientsdata
    clientsdata = []
    headers['Authorization'] = f'{bearer}, {token}'
    totalcount = -1
    while totalcount == -1 or page <= totalcount/200 + 1:
        url = f"https://api.yclients.com/api/v1/company/{salon}/clients/search"
        payload = json.dumps({
            "page": page,
            "page_size": 200,
            "fields": [
                "id",
                "phone",
                "name",
                "surname",
                "patronymic"
            ]
        })
        response = requests.request("POST", url, headers=headers, data=payload).json()
        if totalcount == -1:
            totalcount = response["meta"]["total_count"]
        for item in response["data"]:
            clientsdata.append({
                "id": item.get("id", ""),
                "phone" : item.get("phone", ""),
                "name": item.get("name", ""),
                "patronymic": item.get("patronymic", ""),
                "surname": item.get("surname", "")
            })
        page += 1
    safeClients(salon,clientsdata)
    return clientsdata

@fiosplitter.route('/parsefio', methods=['POST'])
def parsefio():
    try:
        global newdata
        response = request.json
        salon = response["salon_id"]
        login = response["login"]
        password = response["password"]
        newdata = parseclients(clientsdata)
        print(newdata)
        return jsonify({'status': 'success',  'text': newdata})
    except Exception as e:
        return jsonify({'status': 'error', 'text': f'{e}'})
def parseclients(clients):
    parser = NamesParser()
    for item in clients:
        parsename = parser.parse(f'{item["surname"]} {item["name"]} {item["patronymic"]}')
        if parsename["parsed"] == True:
            item["name"] = parsename.get("fn","")
            item["surname"] = parsename.get("sn", "")
            item["patronymic"] = parsename.get("mn", "")
        if parsename["parsed"] == False:
            item["error"] = True
    return clients


@fiosplitter.route('/getError', methods=['GET'])
def getError():
    try:
        global newdata
        global errors
        errors = []
        for item in newdata:
            if 'error' in item:
                errors.append(item)
        return jsonify({'status': 'success',  'text': errors})
    except Exception as e:
        return jsonify({'status': 'error', 'text': f'{e}'})

@fiosplitter.route('/saveClients', methods=['POST'])
def saveClients():
    try:
        response = request.json
        salon = response["salon_id"]
        login = response["login"]
        password = response["password"]
        if password == "":
            token = login
            saveResult(salon, newdata, headers, token)
        else:
            usertoken = getToken(login,password)
            saveResult(salon, newdata, headers, usertoken)

        return jsonify({'status': 'success',  'text': newdata})
    except Exception as e:
        return jsonify({'status': 'error', 'text': f'{e}'})


def saveResult(salon, data, headers, usertoken):
    headers['Authorization'] = f'{bearer}, {usertoken}'
    for item in data:
        url = f'https://api.yclients.com/api/v1/client/{salon}/{item["id"]}'
        payload = json.dumps({
            "id": item['id'],
            "name": item['name'],
            "patronymic": item['patronymic'],
            "phone": item['phone'],
            "surname": item['surname']
        })
        print(url)
        response = requests.request("PUT", url, headers=headers, data=payload)
        print(response.text)

@fiosplitter.route('/getReport')
def get_report():
    global errors

    # Создаем новую workbook и worksheet в Excel
    workbook = Workbook()
    worksheet = workbook.active

    # Записываем заголовки
    headers = ['id', 'Телефон', 'Имя', 'Отчество', 'Фамилия']
    for col, header in enumerate(headers, start=1):
        worksheet.cell(row=1, column=col, value=header)

    # Записываем данные
    for row, item in enumerate(errors, start=2):
        worksheet.cell(row=row, column=1, value=item['id'])
        worksheet.cell(row=row, column=2, value=item['phone'])
        worksheet.cell(row=row, column=3, value=item['name'])
        worksheet.cell(row=row, column=4, value=item['patronymic'])
        worksheet.cell(row=row, column=5, value=item['surname'])

    # Создаем временный буфер для файла
    output = io.BytesIO()

    # Сохраняем workbook во временный буфер
    workbook.save(output)

    # Закрываем workbook
    workbook.close()

    # Подготавливаем response с файлом
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.ms-excel'
    response.headers['Content-Disposition'] = 'attachment; filename=report.xlsx'

    return response

